"""
Importación única de datos desde 'LISTADO ACCESO EMPRESAS 2026.xlsx' (hoja EMPRESAS)
hacia la base de datos de GestorPass.

Uso:
    python scripts/import_excel.py
    python scripts/import_excel.py --apply      (sin --apply solo hace un dry-run)
"""
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                          'LISTADO ACCESO EMPRESAS 2026.xlsx')

COMPANY_PATTERN = re.compile(
    r'\b(S\.?A\.?S\.?|LTDA|S\.?A\.?|COOP|UT\b|UNI[OÓ]N TEMPORAL|FUNDACI[OÓ]N|'
    r'CORPORACI[OÓ]N|CLUB|JUNTA|ASOCIACI[OÓ]N|GRUPO EMPRESARIAL|COMPA[ÑN][IÍ]A|'
    r'SOCIEDAD|CONCEJO)\b', re.IGNORECASE)

EMAIL_DOMAIN_SITES = {
    'gmail.com': 'Gmail',
    'hotmail.com': 'Outlook / Hotmail',
    'hotmail.es': 'Outlook / Hotmail',
    'outlook.com': 'Outlook / Hotmail',
    'outlook.es': 'Outlook / Hotmail',
    'live.com': 'Outlook / Hotmail',
    'yahoo.com': 'Yahoo Mail',
    'yahoo.es': 'Yahoo Mail',
}
FALLBACK_EMAIL_SITE = 'Correo Electrónico'


def clean_text(v):
    if v is None:
        return ''
    return str(v).replace('\xa0', ' ').strip()


def digits_only(v):
    return re.sub(r'\D', '', str(v) if v is not None else '')


def guess_entity_type(name):
    return 'empresa' if COMPANY_PATTERN.search(name) else 'persona'


def get_or_create_site(db, Site, name, category, url=''):
    site = Site.query.filter_by(name=name).first()
    if not site:
        site = Site(name=name, url=url, category=category)
        db.session.add(site)
        db.session.flush()
    return site


def parse_correo(text):
    """Devuelve lista de (email, password) a partir del texto multilinea."""
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    pairs = []
    i = 0
    while i < len(lines):
        if '@' in lines[i]:
            email = lines[i]
            password = ''
            if i + 1 < len(lines) and '@' not in lines[i + 1]:
                password = lines[i + 1]
                i += 2
            else:
                i += 1
            pairs.append((email, password))
        else:
            i += 1
    return pairs


def parse_user_pass(text):
    """Busca patrones 'usuario: X ... contraseña: Y' en texto libre."""
    u = re.search(r'usuario\s*:?\s*([^\s]+)', text, re.IGNORECASE)
    p = re.search(r'contrase\w*\s*:?\s*([^\s]+)', text, re.IGNORECASE)
    username = u.group(1) if u else ''
    password = p.group(1) if p else ''
    return username, password


def parse_camara(text):
    correo = re.search(r'correo\s*:?\s*(\S+@\S+)', text, re.IGNORECASE)
    cc = re.search(r'\bCC\b\s*:?\s*(\d+)', text, re.IGNORECASE)
    clave = re.search(r'clave\s*:?\s*(\S+)', text, re.IGNORECASE)
    return (correo.group(1) if correo else '',
            cc.group(1) if cc else '',
            clave.group(1) if clave else '')


def main(apply_changes):
    from app import create_app, db
    from app.models import Entity, Site, Credential, User

    app = create_app()
    with app.app_context():
        admin = User.query.filter_by(role='admin').first()
        if not admin:
            print('No se encontró un usuario admin. Abortando.')
            return

        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb['EMPRESAS']

        stats = {'entidades_nuevas': 0, 'entidades_actualizadas': 0,
                 'credenciales_nuevas': 0, 'filas_omitidas': 0}

        for r in range(3, ws.max_row + 1):
            razon = clean_text(ws.cell(row=r, column=3).value)
            nit_raw = ws.cell(row=r, column=4).value
            nit_digits = digits_only(nit_raw)

            if not razon or len(nit_digits) < 5:
                stats['filas_omitidas'] += 1
                continue

            lugar = clean_text(ws.cell(row=r, column=2).value).replace('\n', ', ')
            rep_legal = clean_text(ws.cell(row=r, column=5).value)
            cc_rl = clean_text(ws.cell(row=r, column=6).value)
            correo_raw = clean_text(ws.cell(row=r, column=7).value)
            dian_raw = clean_text(ws.cell(row=r, column=8).value)
            aportes_raw = clean_text(ws.cell(row=r, column=10).value)
            camara_raw = clean_text(ws.cell(row=r, column=11).value)
            obligaciones_raw = clean_text(ws.cell(row=r, column=12).value)

            notes_parts = []
            if rep_legal:
                notes_parts.append(f'Representante legal: {rep_legal}' +
                                   (f' (CC {cc_rl})' if cc_rl else ''))
            if obligaciones_raw:
                notes_parts.append('Obligaciones: ' + obligaciones_raw.replace('\n', ', '))
            notes = '\n'.join(notes_parts)

            entity = Entity.query.filter_by(cedula_nit=nit_digits).first()
            if entity:
                entity.place = lugar or entity.place
                if notes:
                    entity.notes = notes
                stats['entidades_actualizadas'] += 1
            else:
                entity = Entity(
                    cedula_nit=nit_digits,
                    name=razon,
                    entity_type=guess_entity_type(razon),
                    place=lugar,
                    notes=notes,
                    created_by_id=admin.id,
                )
                db.session.add(entity)
                db.session.flush()
                stats['entidades_nuevas'] += 1

            # ── Correos ──────────────────────────────────────────────
            if correo_raw:
                pairs = parse_correo(correo_raw)
                if pairs:
                    for email, password in pairs:
                        domain = email.split('@')[-1].lower()
                        site_name = EMAIL_DOMAIN_SITES.get(domain, FALLBACK_EMAIL_SITE)
                        site = get_or_create_site(db, Site, site_name, 'correo')
                        existing = Credential.query.filter_by(
                            entity_id=entity.id, site_id=site.id, username=email).first()
                        if not existing:
                            cred = Credential(entity_id=entity.id, site_id=site.id,
                                              username=email, created_by_id=admin.id)
                            if password:
                                cred.set_password(password)
                            db.session.add(cred)
                            stats['credenciales_nuevas'] += 1
                else:
                    entity.notes = (entity.notes + '\n' if entity.notes else '') + \
                                   f'Correo (sin procesar): {correo_raw}'

            # ── DIAN ─────────────────────────────────────────────────
            if dian_raw:
                lines = [l.strip() for l in dian_raw.split('\n') if l.strip()]
                password = lines[0] if lines else dian_raw
                extra = '\n'.join(lines[1:]) if len(lines) > 1 else ''
                site = get_or_create_site(db, Site, 'DIAN', 'impuestos',
                                          'https://www.dian.gov.co')
                existing = Credential.query.filter_by(entity_id=entity.id, site_id=site.id).first()
                if not existing:
                    cred = Credential(entity_id=entity.id, site_id=site.id,
                                      username=nit_digits, notes=extra,
                                      created_by_id=admin.id)
                    cred.set_password(password)
                    db.session.add(cred)
                    stats['credenciales_nuevas'] += 1

            # ── Aportes en Línea ─────────────────────────────────────
            if aportes_raw:
                username, password = parse_user_pass(aportes_raw)
                site = get_or_create_site(db, Site, 'Aportes en Línea', 'aportes',
                                          'https://www.aportesenlinea.com/independientes/inicio')
                existing = Credential.query.filter_by(entity_id=entity.id, site_id=site.id).first()
                if not existing:
                    cred = Credential(entity_id=entity.id, site_id=site.id,
                                      username=username, notes=aportes_raw,
                                      created_by_id=admin.id)
                    if password:
                        cred.set_password(password)
                    db.session.add(cred)
                    stats['credenciales_nuevas'] += 1

            # ── Cámara de Comercio ───────────────────────────────────
            if camara_raw:
                correo, cc, clave = parse_camara(camara_raw)
                site = get_or_create_site(db, Site, 'Cámara de Comercio CCB', 'camara',
                                          'https://www.ccb.org.co')
                existing = Credential.query.filter_by(entity_id=entity.id, site_id=site.id).first()
                if not existing:
                    cred = Credential(entity_id=entity.id, site_id=site.id,
                                      username=correo, alt_username=cc, notes=camara_raw,
                                      created_by_id=admin.id)
                    if clave:
                        cred.set_password(clave)
                    db.session.add(cred)
                    stats['credenciales_nuevas'] += 1

        print('Resumen de importación:', stats)

        if apply_changes:
            db.session.commit()
            print('Cambios guardados en la base de datos.')
        else:
            db.session.rollback()
            print('DRY-RUN: no se guardó nada. Ejecuta con --apply para confirmar.')


if __name__ == '__main__':
    main(apply_changes='--apply' in sys.argv)
